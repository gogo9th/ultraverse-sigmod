import openpyxl
import datetime
import os
import sys

RESULT_PATH = None


def set_result_path(path):
    global RESULT_PATH
    RESULT_PATH = path


def eprint(*args, **kwargs):
    if RESULT_PATH is None:
        print(*args, file=sys.stderr, **kwargs)
    else:
        print(*args, file=sys.stderr, **kwargs)
        with open(RESULT_PATH, mode='a+') as f:
            f.write(args[0])
            f.write('\n')


def get_elapsed_ms(st, ed):
    if st is None or ed is None:
        return None
    s_time = datetime.datetime.strptime(st, '%Y-%m-%d %H:%M:%S.%f')
    e_time = datetime.datetime.strptime(ed, '%Y-%m-%d %H:%M:%S.%f')
    return '%.0f' % ((e_time - s_time).total_seconds() * 1000)


def get_elapsed_s(st, ed):
    if st is None or ed is None:
        return None
    s_time = datetime.datetime.strptime(st, '%Y-%m-%d %H:%M:%S.%f')
    e_time = datetime.datetime.strptime(ed, '%Y-%m-%d %H:%M:%S.%f')
    return '%.3f' % ((e_time - s_time).total_seconds())


def get_value_script(line, target):
    target = target.strip()
    line = line.strip()
    pos = line.find(target)
    if pos <= 0:
        return None
    if line[pos:] != target:
        return None

    v = line[:pos].strip()
    return v[1:-1]


def read_log(path, keyword):
    def get_value(line, target):
        line = line.strip()
        pos = line.find(target)
        if pos != 0:
            return None

        v = line[pos + len(target):].strip()
        return v

    q1 = 0
    q2 = 0

    with open(path, mode='rt', errors='ignore') as f:
        for line in f:
            if line.find('Query\t') > 0:
                q1 += 1

    start_count = False
    with open(path, mode='rt', errors='ignore') as f:
        for line in f:
            if start_count is True:
                if line.find('Query\t' + keyword) > 0:
                    q2 += 1
                    continue

            if line.find('SELECT * INTO OUTFILE') > 0:
                start_count = True

    eprint('total query count (from general log) : %d' % q1)
    eprint('undo & redo query count (from general log) : %d' % q2)

    return q1, q2


def read_flash(path):
    get_value = get_value_script

    t1 = None
    t2 = None
    t3 = None
    t4 = None

    prepare_st = None
    prepare_ed = None
    update_st = None
    update_ed = None

    with open(path, mode='rt', errors='ignore') as f:
        for line in f:
            v = get_value_script(line, 'prepare data...')
            if v is not None:
                prepare_st = v
                continue
            v = get_value_script(line, 'prepare data...done')
            if v is not None:
                prepare_ed = v
                continue
            v = get_value_script(line, 'update data...')
            if v is not None:
                update_st = v
                continue
            v = get_value_script(line, 'update data...done')
            if v is not None:
                update_ed = v
                continue
            v = get_value(line, 'run flashback undo...done')
            if v is not None:
                t2 = v
                continue
            v = get_value(line, 'apply user query...done')
            if v is not None:
                t3 = v
                continue
            v = get_value(line, 'run flashback redo...done')
            if v is not None:
                t4 = v
                continue
            v = get_value(line, 'run flashback undo...')
            if v is not None:
                t1 = v
                continue

    return t1, t2, t3, t4, get_elapsed_ms(t1, t4), \
        get_elapsed_ms(prepare_st, prepare_ed), get_elapsed_ms(
            update_st, update_ed)


def read_db_state(path, is_django):
    def get_value(line, target):
        line = line.strip()
        pos = line.find(target)
        if pos < 0:
            return None

        v = line[pos + len(target):].strip()

        msec_pos = str(v).rfind('.')
        if msec_pos > 0:
            msec = v[msec_pos + 1:]
            if len(msec) != 6:
                v = v[:msec_pos + 1] + msec.zfill(6)

        return v

    t1 = None
    t2 = None
    t3 = None
    t4 = None
    t5 = None
    t6 = None
    t7 = None
    t8 = None
    t9 = None

    prepare_st = None
    prepare_ed = None
    update_st = None
    update_ed = None
    hash_matched = None

    with open(path, mode='rt', errors='ignore') as f:
        for line in f:
            v = get_value_script(line, 'prepare data...')
            if v is not None:
                prepare_st = v
                continue
            v = get_value_script(line, 'prepare data...done')
            if v is not None:
                prepare_ed = v
                continue
            v = get_value_script(line, 'update data...')
            if v is not None:
                update_st = v
                continue
            v = get_value_script(line, 'update data...done')
            if v is not None:
                update_ed = v
                continue
            v = get_value(line, 'match send event :')
            if v is not None:
                hash_matched = v
                continue
            v = get_value(line, 'start time :')
            if v is not None:
                t1 = v
                continue
            v = get_value(line, 'load log end time :')
            if v is not None:
                t2 = v
                continue
            v = get_value(line, 'query analyze end time :')
            if v is not None:
                t3 = v
                continue
            v = get_value(line, 'undo end time :')
            if v is not None:
                t4 = v
                continue
            v = get_value(line, 'redo end time :')
            if v is not None:
                t5 = v
                continue
            v = get_value(line, 'sync table end time :')
            if v is not None:
                t6 = v
                continue
            v = get_value(line, 'swap table end time :')
            if v is not None:
                t7 = v
                continue
            v = get_value(line, 'end time :')
            if v is not None:
                t8 = v
                continue
            v = get_value(line, 'elapsed time :')
            if v is not None:
                pos = v.find('ms')
                if pos >= 0:
                    t9 = v[0:pos].strip()
                else:
                    t9 = v
                continue

    t9 = '%.3f' % (float(t9) / 1000)

    if is_django:
        return t1, t2, t3, t4, t5, t6, t7, t8, t9, \
            get_elapsed_s(t1, t2), \
            get_elapsed_s(t2, t3), \
            get_elapsed_s(t3, t4), \
            get_elapsed_s(t4, t5), \
            get_elapsed_s(t5, t6), \
            get_elapsed_s(t6, t7)

    else:
        eprint('start time : %s' % t1)
        eprint('load log end time  : %s' % t2)
        eprint('query analyze end time  : %s' % t3)
        eprint('undo end time  : %s' % t4)
        eprint('redo end time  : %s' % t5)
        eprint('sync table end time  : %s' % t6)
        eprint('swap table end time  : %s' % t7)
        eprint('hash matched time  : %s' % hash_matched)
        eprint('end time  : %s' % t8)
        eprint('elapsed time(s) : %s' % t9)
        eprint('db prepare elapsed time(s) : %s' %
               get_elapsed_s(prepare_st, prepare_ed))
        eprint('update data elapsed time(s) : %s' %
               get_elapsed_s(update_st, update_ed))
        eprint('load log elapsed time (s) : %s' % get_elapsed_s(t1, t2))
        eprint('query analyze elapsed time (s) : %s' % get_elapsed_s(t2, t3))
        eprint('undo elapsed time (s) : %s' % get_elapsed_s(t3, t4))
        eprint('redo elapsed time (s) : %s' % get_elapsed_s(t4, t5))
        eprint('hash matched elapsed time (s) : %s' %
               get_elapsed_s(t4, hash_matched))
        eprint('sync table elapsed time (s) : %s' % get_elapsed_s(t5, t6))
        eprint('swap table elapsed time (s) : %s' % get_elapsed_s(t6, t7))

        return t1, t2, t3, t4, t5, t6, t7, t8, t9, \
            get_elapsed_s(prepare_st, prepare_ed), get_elapsed_s(update_st, update_ed), \
            get_elapsed_s(t1, t2), \
            get_elapsed_s(t2, t3), \
            get_elapsed_s(t3, t4), \
            get_elapsed_s(t4, t5), \
            get_elapsed_s(t5, t6), \
            get_elapsed_s(t6, t7)


if __name__ == '__main__':
    is_django = True
    base_dir = 'D:\\PJT\\OLD\\DB state change\\전달\\15차 전달(django)'

    sub_dir = [d for d in os.listdir(
        base_dir) if os.path.isdir(os.path.join(base_dir, d))]

    out_wb = openpyxl.Workbook()
    ws = out_wb.get_active_sheet()

    col_idx = 1
    for i in range(len(sub_dir)):
        row_idx = 1

        print('%d / %d (%s)' % (i + 1, len(sub_dir), sub_dir[i]))
        ws.cell(row=row_idx, column=col_idx).value = sub_dir[i]
        row_idx += 1

        if os.path.exists(os.path.join(base_dir, sub_dir[i], 'db_state.txt')) is True:
            t = read_db_state(os.path.join(
                base_dir, sub_dir[i], 'db_state.txt'), is_django)
            for j in range(0, len(t)):
                ws.cell(row=(j + row_idx), column=col_idx).value = t[j]
            row_idx += j + 1

            t = read_log(os.path.join(
                base_dir, sub_dir[i], 'general_modified.log'), '/* GENERATED BY STATE QUERY */')
            for j in range(0, len(t)):
                ws.cell(row=(j + row_idx), column=col_idx).value = t[j]

        elif os.path.exists(os.path.join(base_dir, sub_dir[i], 'flash.txt')) is True:
            t = read_db_state(os.path.join(
                base_dir, sub_dir[i], 'flash.txt'), is_django)
            for j in range(0, len(t)):
                ws.cell(row=(j + row_idx), column=col_idx).value = t[j]
            row_idx += j + 1

            t = read_log(os.path.join(
                base_dir, sub_dir[i], 'general_origin.log'), '/* GENERATED BY STATE QUERY */')
            for j in range(0, len(t)):
                ws.cell(row=(j + row_idx), column=col_idx).value = t[j]

        col_idx += 1

    out_wb.save('summary_tmp.xlsx')
    out_wb.close()
